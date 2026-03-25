import streamlit as st
import re
import io
import zipfile
import pdfplumber
import uuid
import pandas as pd
import base64  # ⭐️ 이미지 변환을 위해 추가됨
from pypdf import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- 페이지 기본 설정 ---
# [공통 규칙 2] 레이아웃 "centered" 적용
st.set_page_config(page_title="매출전표 정리 자동화", page_icon="🧾", layout="centered")

# ==========================================
# 🔒 보안(비밀번호) 설정 영역 (폼 형태)
# [공통 규칙 3] 페이지 설정 직후 비밀번호 로직 처리 및 st.stop() 적용
# ==========================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.warning("🔒 보안을 위해 비밀번호를 입력해주세요.")
    with st.form("login_form"):
        pwd = st.text_input("비밀번호", type="password")
        submitted = st.form_submit_button("확인")
        
        if submitted:
            # st.secrets에 값이 없으면 기본값(ip2b)으로 작동하도록 방어 코드를 넣었습니다.
            expected_pwd = st.secrets.get("APP_PASSWORD", "ip2b") 
            if pwd == expected_pwd:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("비밀번호가 일치하지 않습니다.")
    st.stop() # 비밀번호 통과 전까지 아래 코드 실행 중지
# ==========================================

# --- 🎨 통합 CSS (플로팅 로고 및 파일 목록 스타일) ---
# [공통 규칙 6] UI 최적화 (불필요한 안내 문구 숨기기)
st.markdown("""
    <style>
    /* 'Press Enter to submit form' 안내 문구 강제로 숨기기 */
    div[data-testid="InputInstructions"] {
        display: none !important;
    }

    /* 업로드된 파일 목록 카드 스타일 */
    [data-testid="stUploadedFile"] {
        background-color: #f8f9fa;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        padding: 5px 10px;
        margin-top: 8px;
    }
    
    /* [공통 규칙 4] 우측 상단 플로팅 로고 스타일 및 모바일 대응 */
    .company-logo {
        position: fixed;
        top: 70px;      
        right: 30px;    
        width: 110px;   
        z-index: 1000;  
        cursor: pointer;
    }
    @media (max-width: 640px) {
        .company-logo {
            width: 80px;
            top: 60px;
            right: 10px; 
        }
    }
    </style>
""", unsafe_allow_html=True)

# --- 🖼️ 로고 변환 함수 ---
def get_base64_of_bin_file(bin_file):
    try:
        with open(bin_file, 'rb') as f:
            return base64.b64encode(f.read()).decode()
    except:
        return "" 

# --- 🖼️ 제작사 로고 화면에 띄우기 (HTML) ---
# 주의: 같은 폴더 안에 company_logo.png 파일이 있어야 합니다.
comp_img_base64 = get_base64_of_bin_file("company_logo.png") 

if comp_img_base64:
    st.markdown(
        f"""
        <a href="http://www.iptob.co.kr/" target="_blank" title="(주)아이피투비 홈페이지로 이동">
            <img src="data:image/png;base64,{comp_img_base64}" class="company-logo" alt="(주)아이피투비 로고">
        </a>
        """,
        unsafe_allow_html=True
    )

# --- 사이드바: 홈 버튼 및 얇은 여백 구분선 ---
# [공통 규칙 5] 사이드바 최상단 홈 버튼 배치 (두꺼운 기본 버튼 대신 텍스트 링크 사용)
with st.sidebar:
    st.markdown(
        '''
        <div style="margin-top: 5px;">
            <a href="https://ip2b-work-tools.streamlit.app/" target="_blank" style="text-decoration: none; color: #31333F; font-size: 15px; font-weight: 600;">
                🏠 홈으로
            </a>
        </div>
        <hr style="margin-top: 10px; margin-bottom: 15px; border: 0; border-top: 1px solid rgba(49, 51, 63, 0.2);">
        ''', 
        unsafe_allow_html=True
    )

# --- 세션 상태(Session State) 초기화 ---
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = str(uuid.uuid4())
if "is_processed" not in st.session_state:
    st.session_state.is_processed = False
if "zip_data" not in st.session_state:
    st.session_state.zip_data = None
if "stats" not in st.session_state:
    st.session_state.stats = {"total": 0, "success": 0, "error": 0}
if "preview_data" not in st.session_state:
    st.session_state.preview_data = []

def reset_app():
    st.session_state.uploader_key = str(uuid.uuid4())
    st.session_state.is_processed = False
    st.session_state.zip_data = None
    st.session_state.stats = {"total": 0, "success": 0, "error": 0}
    st.session_state.preview_data = []

# --- 데이터 추출 로직 함수 ---
def extract_receipt_info(text_layout, text_normal):
    date_str, time_str, store_name, supply_val, vat_val, card_str, total_amount = ("", "", "", "", "", "", "")

    card_match = re.search(r'\*{4}\s*-\s*(\d{4})', text_normal)
    if card_match:
        card_str = card_match.group(1)

    layout_date = re.search(r'(\d{4})년\s*(\d{2})월\s*(\d{2})일', text_layout)
    layout_time = re.search(r'(\d{2})시\s*(\d{2})분', text_layout)
    layout_store = re.search(r'가맹점명\s+(.+)', text_layout)
    layout_supply = re.search(r'공급가액\s+([\d,\.]+)', text_layout)
    layout_vat = re.search(r'부가세\s+([\d,\.]+)', text_layout)
    layout_total = re.search(r'총액\s+([\d,\.]+)', text_layout)

    if layout_date and layout_store and layout_total:
        date_str = f"{layout_date.group(1)}-{layout_date.group(2)}-{layout_date.group(3)}"
        time_str = f"{layout_time.group(1)}:{layout_time.group(2)}" if layout_time else ""
        store_name = re.sub(r'[\\/*?:"<>|]', "", layout_store.group(1).strip())
        supply_val = re.sub(r'[^\d]', '', layout_supply.group(1)) if layout_supply else ""
        vat_val = re.sub(r'[^\d]', '', layout_vat.group(1)) if layout_vat else ""
        total_amount = layout_total.group(1).strip()
    else:
        fallback = re.search(r'(\d{4})년\s*(\d{2})월\s*(\d{2})일\s*(\d{2})시\s*(\d{2})분.*?\n(?:.*?\n)?([\d,\.]+)원\n([\d,\.]+)원\n.*?\n([\d,\.]+)원\n\d{8}\n([^\n]+)', text_normal)
        if fallback:
            date_str = f"{fallback.group(1)}-{fallback.group(2)}-{fallback.group(3)}"
            time_str = f"{fallback.group(4)}:{fallback.group(5)}"
            supply_val = re.sub(r'[^\d]', '', fallback.group(6))
            vat_val = re.sub(r'[^\d]', '', fallback.group(7))
            total_amount = fallback.group(8).strip()
            store_name = re.sub(r'[\\/*?:"<>|]', "", fallback.group(9).strip())
            
    return date_str, time_str, store_name, supply_val, vat_val, card_str, total_amount

# --- 메인 화면 ---
st.title("🧾 매출전표 정리 자동화")

with st.expander("💡 사용 방법"):
    st.write("""
    1. 스캔된 PDF 원본 파일들을 아래 영역에 마우스로 끌어다 놓습니다.
    2. **[작업 실행]** 버튼을 클릭합니다.
    3. 처리가 완료되면 화면 중앙에서 **추출된 데이터를 미리 확인**할 수 있습니다.
    4. 좌측 **사이드바**에서 **[다운로드]** 버튼을 눌러 압축(ZIP) 파일을 받습니다.
    """)

uploaded_files = st.file_uploader(
    "매출전표 PDF 파일을 업로드해 주세요.", 
    type=['pdf'], 
    accept_multiple_files=True, 
    key=st.session_state.uploader_key
)

# --- 데이터 처리 로직 ---
if uploaded_files and not st.session_state.is_processed:
    
    st.success(f"✅ 총 **{len(uploaded_files)}개**의 파일이 성공적으로 업로드되었습니다. 아래 버튼을 눌러주세요.")
    
    # [공통 규칙 7] 여백이 얇은 구분선 활용 (st.divider 교체)
    st.markdown('<hr style="margin-top: 15px; margin-bottom: 15px; border: 0; border-top: 1px solid rgba(49, 51, 63, 0.2);">', unsafe_allow_html=True)
    
    if st.button("작업 실행", type="primary", use_container_width=True):
        
        progress_text = st.empty()
        progress_bar = st.progress(0)
        
        total_files = len(uploaded_files)
        success_pages = 0
        error_pages = 0
        
        zip_buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active 
        ws.title = "지출결의내역"
        
        headers = ["A(빈칸)", "결제일", "C(빈칸)", "시간", "가맹점명", "F(빈칸)", "공급가액", "부가세액", "I(빈칸)", "J(빈칸)", "카드번호(끝4자리)"]
        ws.append(headers)

        header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        seen_filenames = set()
        preview_list = [] 

        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for file_idx, uploaded_file in enumerate(uploaded_files):
                progress_text.markdown(f"⏳ **데이터 추출 중 ({file_idx + 1}/{total_files}):** `{uploaded_file.name}`")
                progress_bar.progress((file_idx + 1) / total_files)
                
                file_bytes = uploaded_file.read()
                try:
                    reader = PdfReader(io.BytesIO(file_bytes))
                except Exception:
                    st.error(f"❌ 파일 읽기 오류: {uploaded_file.name}")
                    continue

                with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                    for i, page in enumerate(pdf.pages):
                        text_layout = page.extract_text(layout=True) or ""
                        text_normal = page.extract_text() or ""
                        
                        date_str, time_str, store_name, supply_val, vat_val, card_str, total_amount = extract_receipt_info(text_layout, text_normal)

                        writer = PdfWriter()
                        writer.add_page(reader.pages[i])
                        pdf_out_buffer = io.BytesIO()
                        writer.write(pdf_out_buffer)

                        if date_str:
                            y, m, d = date_str.split("-")
                            year_short = y[2:]
                            base_filename = f"{year_short}.{m}.{d}_{store_name}_{total_amount}"
                            
                            new_filename_with_ext = f"{base_filename}.pdf"
                            counter = 1
                            while new_filename_with_ext in seen_filenames:
                                new_filename_with_ext = f"{base_filename}_{counter}.pdf"
                                counter += 1
                            seen_filenames.add(new_filename_with_ext)

                            zip_file.writestr(new_filename_with_ext, pdf_out_buffer.getvalue())
                            
                            num_supply = int(supply_val) if supply_val.isdigit() else 0
                            num_vat = int(vat_val) if vat_val.isdigit() else 0
                            num_card = int(card_str) if card_str.isdigit() else "" 

                            excel_row = ["", date_str, "", time_str, store_name, "", num_supply, num_vat, "", "", num_card]
                            ws.append(excel_row)
                            
                            current_row = ws.max_row
                            ws[f'G{current_row}'].number_format = '#,##0'
                            ws[f'H{current_row}'].number_format = '#,##0'
                            ws[f'K{current_row}'].number_format = '0000'
                            
                            success_pages += 1
                            
                            preview_list.append({
                                "결제일": date_str,
                                "시간": time_str,
                                "가맹점명": store_name,
                                "공급가액": num_supply,
                                "부가세액": num_vat,
                                "총액": total_amount,
                                "카드번호": f"*{num_card}" if num_card else "",
                                "원본 파일명": f"{uploaded_file.name} ({i+1}p)"
                            })
                            
                        else:
                            error_filename = f"검토요망_{uploaded_file.name}_page{i+1}.pdf"
                            zip_file.writestr(f"검토요망/{error_filename}", pdf_out_buffer.getvalue())
                            error_pages += 1

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter 
                for cell in col:
                    if cell.value:
                        cell_len = sum(2 if ord(c) > 127 else 1.2 for c in str(cell.value))
                        if len(str(cell.value)) > 0 and cell_len > max_length:
                            max_length = cell_len
                
                adjusted_width = (max_length + 3)
                ws.column_dimensions[column].width = adjusted_width

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("매출전표내역.xlsx", excel_buffer.getvalue())
            
        st.session_state.zip_data = zip_buffer.getvalue()
        st.session_state.stats = {
            "total": total_files,
            "success": success_pages,
            "error": error_pages
        }
        st.session_state.preview_data = preview_list
        st.session_state.is_processed = True 
        
        st.rerun()

# --- 메인 화면 데이터 미리보기 표 ---
if st.session_state.is_processed:
    if st.session_state.preview_data:
        st.subheader("🔎 미리보기")
        df = pd.DataFrame(st.session_state.preview_data)
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("정상적으로 추출된 데이터가 없습니다. (모두 검토 요망으로 분류됨)")

# --- 사이드바 결과 화면 ---
if st.session_state.is_processed:
    st.success("✨ 작업이 완료되었습니다! 데이터를 확인하고 다운로드해 주세요.")
    
    with st.sidebar:
        st.header("📊 처리 결과")
        st.metric("📜 원본 문서", f"{st.session_state.stats['total']} 개")
        st.metric("✅ 정상 처리", f"{st.session_state.stats['success']} 건")
        st.metric("⚠️ 검토 요망", f"{st.session_state.stats['error']} 건")
        
        # [공통 규칙 7] 여백이 얇은 구분선 활용 (st.divider 교체)
        st.markdown('<hr style="margin-top: 15px; margin-bottom: 15px; border: 0; border-top: 1px solid rgba(49, 51, 63, 0.2);">', unsafe_allow_html=True)
        
        st.download_button(
            label="📦 다운로드 (ZIP)",
            data=st.session_state.zip_data,
            file_name="매출전표_완료.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True
        )
        
        st.button("🔄 초기화", on_click=reset_app, use_container_width=True)
